# -*- coding: utf-8 -*-
"""
Генератор xlsx-трекера для ручной публикации объявлений на Авито.

Создаёт книгу с двумя листами:
  - "Трекер объявлений" — таблица с данными + дропдаун статуса + форматирование
  - "Инструкция"        — расшифровка колонок и статусов

Использование:
    py -3 -m pip install openpyxl --quiet      (один раз)
    py -3 06-tracker-generator.py

Перед запуском заполните блок "ДАННЫЕ" — список строк под свою тематику
(можно сгенерировать программно из своей матрицы/папки листингов вместо
ручного списка ROWS).

Зависимости: openpyxl
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ───────────────────────── НАСТРОЙКИ ─────────────────────────
OUTPUT_PATH = Path(__file__).parent.parent / "docs" / "avito-tracker.xlsx"

COLUMNS = [
    # (заголовок,            ширина)
    ("№",                    5),
    ("Заголовок",            48),
    ("Направление",          18),
    ("Материал",             22),
    ("Целевой ключ",         42),
    ("Цена",                 26),
    ("Статус",               18),
    ("Дата публ.",           14),
    ("Ссылка Авито",         32),
    ("Просмотры",            13),
    ("Контакты",             12),
]

STATUS_COLUMN_INDEX = 7          # 1-based индекс колонки "Статус" в COLUMNS
STATUS_OPTIONS = ["черновик", "на модерации", "опубликовано"]
DEFAULT_STATUS = "черновик"

# ───────────────────────── ДАННЫЕ ─────────────────────────
# Каждая строка — кортеж значений по колонкам, КРОМЕ "№", "Статус", "Дата публ.",
# "Ссылка Авито", "Просмотры", "Контакты" — они заполняются автоматически
# (№ — по порядку, Статус — DEFAULT_STATUS, остальные — пусто, заполняются вручную после публикации).
#
# Формат: (Заголовок, Направление, Материал, Целевой ключ, Цена)
ROWS = [
    # ("Раскрой фанеры на ЧПУ — мебельные детали", "ЧПУ-фрезеровка", "Фанера 10–20 мм",
    #  "Раскрой фанеры на ЧПУ в Москве", "от 36 ₽/п.м, мин. 2 500 ₽"),
    # ... добавьте свои строки по образцу — например, прочитав их из docs/matrix-N.md
]
# ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ROW_FILL_ALT = PatternFill("solid", fgColor="EBF3FB")
STATUS_DRAFT_FILL = PatternFill("solid", fgColor="FFF2CC")


def build_tracker_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "Трекер объявлений"

    for col_idx, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_offset, row_data in enumerate(ROWS):
        row_idx = row_offset + 2
        title, direction, material, keyword, price = row_data
        values = [row_offset + 1, title, direction, material, keyword, price,
                  DEFAULT_STATUS, None, None, None, None]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (2, 5)))
            if row_offset % 2 == 1:
                cell.fill = ROW_FILL_ALT

        status_cell = ws.cell(row=row_idx, column=STATUS_COLUMN_INDEX)
        if status_cell.value == DEFAULT_STATUS:
            status_cell.fill = STATUS_DRAFT_FILL

    last_row = len(ROWS) + 1
    if last_row >= 2:
        dv = DataValidation(type="list", formula1=f'"{",".join(STATUS_OPTIONS)}"', allow_blank=True)
        status_letter = get_column_letter(STATUS_COLUMN_INDEX)
        dv.add(f"{status_letter}2:{status_letter}{last_row}")
        ws.add_data_validation(dv)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(last_row, 1)}"


def build_instructions_sheet(wb: Workbook):
    ws = wb.create_sheet("Инструкция")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    rows = [
        ("Поле", "Как заполнять"),
        ("№", "Порядковый номер — соответствует номеру файла объявления"),
        ("Заголовок", "Название объявления — копируется в заголовок карточки на Авито"),
        ("Направление", "Категория услуги внутри вашей тематики"),
        ("Материал", "Материал/вариант исполнения из строки матрицы"),
        ("Целевой ключ", "Ключевая фраза, под которую оптимизирован заголовок"),
        ("Цена", "Цена/модель из объявления — для контроля при заливке"),
        ("Статус", f"Выбор из списка: {', '.join(STATUS_OPTIONS)} (дропдаун в ячейке)"),
        ("Дата публ.", "Заполняется вручную в момент публикации (формат ДД.ММ.ГГГГ)"),
        ("Ссылка Авито", "Вставить ссылку на опубликованное объявление сразу после появления"),
        ("Просмотры", "Обновлять раз в несколько дней — для оценки эффективности заголовка/фото"),
        ("Контакты", "Количество обращений — для оценки конверсии текста в заявку"),
    ]
    for r, (a, b) in enumerate(rows, start=1):
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        if r == 1:
            for c in (ca, cb):
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
        cb.alignment = Alignment(wrap_text=True, vertical="center")

    ws.freeze_panes = "A2"


def main():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_tracker_sheet(wb)
    build_instructions_sheet(wb)
    wb.save(OUTPUT_PATH)
    print(f"Сохранено: {OUTPUT_PATH}")
    print(f"Строк данных: {len(ROWS)}")
    if not ROWS:
        print("⚠ ROWS пуст — заполните блок ДАННЫЕ перед запуском (например, из своей матрицы).")


if __name__ == "__main__":
    main()
