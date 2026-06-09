# -*- coding: utf-8 -*-
"""
Генератор xlsx-трекера для ручной публикации объявлений на Авито.

Создаёт книгу с двумя листами:
  - "Трекер объявлений" — таблица с данными + дропдаун статуса + форматирование
  - "Инструкция"        — расшифровка колонок и статусов

Использование:
    py -3 tools/generate-tracker.py

Адаптировано из avito-toolkit/06-tracker-generator.py (методология frezorez,
промпт 8) — структура и форматирование книги сохранены без изменений,
изменены только: набор колонок (под матрицу услуг — Группа/Ключ/Пачка вместо
Направление/Материал/Целевой ключ) и источник строк — вместо ручного списка
ROWS читаем уже подготовленный черновик docs/avito-tracker.csv (30 строк из
docs/1-uslugi-matrica-30.md, статус «черновик», пачка 1/2/3 проставлена).

Зависимости: openpyxl
"""

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ───────────────────────── НАСТРОЙКИ ─────────────────────────
PROJECT_ROOT = Path(__file__).parent.parent
SOURCE_CSV = PROJECT_ROOT / "docs" / "avito-tracker.csv"
OUTPUT_PATH = PROJECT_ROOT / "docs" / "avito-tracker.xlsx"

COLUMNS = [
    # (заголовок,        ширина)
    ("№",                5),
    ("Заголовок",        46),
    ("Группа",           14),
    ("Ключ",             30),
    ("Пачка",            8),
    ("Статус",           16),
    ("Дата",             13),
    ("Ссылка",           32),
    ("Просмотры",        12),
    ("Контакты",         11),
    ("Цена",             16),
]

STATUS_COLUMN_INDEX = 6          # 1-based индекс колонки "Статус" в COLUMNS
STATUS_OPTIONS = ["черновик", "на модерации", "опубликовано"]
DEFAULT_STATUS = "черновик"
# ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ROW_FILL_ALT = PatternFill("solid", fgColor="EBF3FB")
STATUS_DRAFT_FILL = PatternFill("solid", fgColor="FFF2CC")


def load_rows() -> list[list]:
    """Читает готовый черновик из CSV (см. docstring) — № и Статус приходят
    готовыми из файла, но переcчитываются/нормализуются на всякий случай."""
    with SOURCE_CSV.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = []
        for i, rec in enumerate(reader, start=1):
            rows.append([
                i,
                rec["Заголовок"],
                rec["Группа"],
                rec["Ключ"],
                int(rec["Пачка"]),
                rec["Статус"] or DEFAULT_STATUS,
                None,   # Дата — заполняется вручную при публикации
                None,   # Ссылка — то же
                None,   # Просмотры
                None,   # Контакты
                rec["Цена"],
            ])
        return rows


def build_tracker_sheet(wb: Workbook, rows: list[list]):
    ws = wb.active
    ws.title = "Трекер объявлений"

    for col_idx, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_offset, values in enumerate(rows):
        row_idx = row_offset + 2
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (2, 4)))
            if row_offset % 2 == 1:
                cell.fill = ROW_FILL_ALT

        status_cell = ws.cell(row=row_idx, column=STATUS_COLUMN_INDEX)
        if status_cell.value == DEFAULT_STATUS:
            status_cell.fill = STATUS_DRAFT_FILL

    last_row = len(rows) + 1
    if last_row >= 2:
        dv = DataValidation(type="list", formula1=f'"{",".join(STATUS_OPTIONS)}"', allow_blank=True)
        status_letter = get_column_letter(STATUS_COLUMN_INDEX)
        dv.add(f"{status_letter}2:{status_letter}{last_row}")
        ws.add_data_validation(dv)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(last_row, 1)}"


def build_instructions_sheet(wb: Workbook):
    ws = wb.create_sheet("Инструкция")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 95

    rows = [
        ("Поле", "Как заполнять"),
        ("№", "Порядковый номер — соответствует номеру файла объявления в prompts/listings/"),
        ("Заголовок", "Название объявления — копируется в заголовок карточки на Авито"),
        ("Группа", "Группа услуги из матрицы (docs/1-uslugi-matrica-30.md): Боты / n8n / Документы / Расчёты / Сайты / Контент"),
        ("Ключ", "Целевой ключевой запрос, под который оптимизирован заголовок и текст"),
        ("Пачка", "Номер пачки публикации (1/2/3) — порядок заливки объявлений, чтобы не публиковать все 30 разом"),
        ("Статус", f"Выбор из списка: {', '.join(STATUS_OPTIONS)} (дропдаун в ячейке)"),
        ("Дата", "Заполняется вручную в момент публикации (формат ДД.ММ.ГГГГ)"),
        ("Ссылка", "Вставить ссылку на опубликованное объявление сразу после появления"),
        ("Просмотры", "Обновлять раз в несколько дней — для оценки эффективности заголовка/обложки карточки"),
        ("Контакты", "Количество обращений — для оценки конверсии текста и цены в заявку"),
        ("Цена", "Цена/модель из текста объявления (вилки по группам — см. docs/research-uslugi-ai.md). "
                  "«договорная» — означает, что итоговую цифру в карточке владелец проставляет сам после сверки с актуальной выдачей Авито."),
    ]
    for r, (a, b) in enumerate(rows, start=1):
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        if r == 1:
            for c in (ca, cb):
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
        cb.alignment = Alignment(wrap_text=True, vertical="center")

    ws.freeze_panes = "A2"


def main():
    rows = load_rows()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_tracker_sheet(wb, rows)
    build_instructions_sheet(wb)
    wb.save(OUTPUT_PATH)
    print(f"Сохранено: {OUTPUT_PATH}")
    print(f"Строк данных: {len(rows)} (источник: {SOURCE_CSV.relative_to(PROJECT_ROOT)})")


if __name__ == "__main__":
    main()
